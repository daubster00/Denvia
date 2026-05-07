"""Story 5.5 — /api/v1/admin/analytics/revenue-variance/export 통합 테스트 (AC-3)."""

from __future__ import annotations

import io
import time
from unittest.mock import AsyncMock, MagicMock, patch

import jwt as pyjwt
import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from api.src.main import app
from api.src.models.base import get_session
from api.src.deps.redis import get_redis_runtime
from api.src.settings import settings


def _admin_jwt() -> str:
    return pyjwt.encode(
        {"sub": "99", "aud": "denvia-admin", "exp": int(time.time()) + 3600},
        settings.denvia_jwt_secret,
        algorithm=settings.denvia_jwt_algorithm,
    )


def _admin_user() -> MagicMock:
    u = MagicMock()
    u.id = 99
    u.role = "admin"
    u.subscription_status = "free"
    u.segment = None
    u.withdrawn_at = None
    u.must_reset_password = False
    return u


def _stub_session():
    s = MagicMock()
    s.execute = AsyncMock()

    async def gen():
        yield s

    return gen


def _stub_redis():
    r = MagicMock()
    r.get = AsyncMock(return_value=None)

    async def gen():
        yield r

    return gen


def _summary(year_month: str = "2026-05"):
    return {
        "year_month": year_month,
        "revenue_krw": 1_485_000,
        "token_cost_usd": "12.345600",
        "token_cost_krw": 17_284,
        "usd_to_krw": 1400,
        "variance_krw": 1_467_716,
        "error_count": 3,
        "anomaly_count": 12,
        "applied_filters": {
            "year_month": year_month,
            "kst_start": f"{year_month}-01T00:00:00+09:00",
            "kst_end_exclusive": "2026-06-01T00:00:00+09:00",
        },
    }


def _detail_row(payment_id: int = 1, **kwargs):
    base = {
        "payment_id": payment_id,
        "charged_at_kst": "2026-05-07 14:23:11",
        "amount_krw": 9900,
        "user_id": 42,
        "email_masked": "u**@example.com",
        "provider_order_id": f"sub-{payment_id}-2026-05-07",
        "subscription_id": 100 + payment_id,
    }
    base.update(kwargs)
    return base


@pytest.fixture(autouse=True)
def clear_overrides():
    yield
    app.dependency_overrides.clear()


@pytest.mark.asyncio
class TestRevenueExportEndpoint:
    async def _call(self, qs: str = ""):
        token = _admin_jwt()
        app.dependency_overrides[get_session] = _stub_session()
        app.dependency_overrides[get_redis_runtime] = _stub_redis()
        with patch(
            "api.src.deps.auth.get_user_by_id",
            new=AsyncMock(return_value=_admin_user()),
        ):
            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                return await client.get(
                    f"/api/v1/admin/analytics/revenue-variance/export{qs}",
                    cookies={"denvia_admin_session": token},
                )

    async def test_year_month_required_422(self):
        res = await self._call()
        assert res.status_code == 422

    async def test_year_month_invalid_format_422(self):
        res = await self._call("?year_month=abc")
        assert res.status_code == 422

    async def test_xlsx_content_type(self):
        rows = [_detail_row(1)]
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=(rows, False)),
        ):
            res = await self._call("?year_month=2026-05")
        assert res.status_code == 200
        assert res.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert 'filename="revenue_variance_2026-05.xlsx"' in res.headers.get(
            "content-disposition", ""
        )
        assert res.headers.get("Cache-Control") == "no-store"

    async def test_summary_sheet_contents(self):
        rows = [_detail_row(1), _detail_row(2)]
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=(rows, False)),
        ):
            res = await self._call("?year_month=2026-05")
        assert res.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        assert "Summary" in wb.sheetnames
        assert "Detail" in wb.sheetnames
        ws = wb["Summary"]
        keys = [row[0].value for row in ws.iter_rows(min_row=2)]
        # 10 항목 이상 노출 (기간 + 매출 + 토큰비용 USD/KRW + 환율 + 차액 + 에러 + 이상 + 행제한 + 잘림)
        assert "당월 매출 (KRW)" in keys
        assert "당월 토큰 비용 (USD)" in keys
        assert "적용 환율 (KRW/USD)" in keys
        assert "당월 토큰 비용 (KRW)" in keys
        assert "차액 (KRW)" in keys
        assert "결제 실패 건수" in keys
        assert "이상 이벤트 건수" in keys
        assert "행 제한 (Detail)" in keys
        assert "잘림 여부" in keys

    async def test_detail_sheet_columns(self):
        rows = [_detail_row(1)]
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=(rows, False)),
        ):
            res = await self._call("?year_month=2026-05")
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb["Detail"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header_row == (
            "payment_id",
            "charged_at_kst",
            "amount_krw",
            "user_id",
            "email_masked",
            "provider_order_id",
            "subscription_id",
        )

    async def test_excel_safe_cell_applied(self):
        # provider_order_id가 '='로 시작하는 악성 입력 — _excel_safe_cell이 prefix `'`을 붙여야 함
        injected = _detail_row(1, provider_order_id="=SUM(A1:A10)")
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=([injected], False)),
        ):
            res = await self._call("?year_month=2026-05")
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb["Detail"]
        # data row = row 2 (header)
        data = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        # provider_order_id = column 6 (1-indexed)
        assert data[5] == "'=SUM(A1:A10)"

    async def test_truncated_header(self):
        rows = [_detail_row(i) for i in range(1, 4)]
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=(rows, True)),
        ):
            res = await self._call("?year_month=2026-05")
        assert res.status_code == 200
        assert res.headers.get("X-Truncated") == "true"

    async def test_email_masked_in_detail(self):
        rows = [_detail_row(1, email_masked="hello**@example.com")]
        with patch(
            "api.src.routers.admin.analytics.get_revenue_variance_month",
            new=AsyncMock(return_value=_summary()),
        ), patch(
            "api.src.routers.admin.analytics.get_revenue_variance_export_rows",
            new=AsyncMock(return_value=(rows, False)),
        ):
            res = await self._call("?year_month=2026-05")
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb["Detail"]
        data = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        assert "**" in str(data[4])

    async def test_unauthenticated_returns_401(self):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            res = await client.get(
                "/api/v1/admin/analytics/revenue-variance/export?year_month=2026-05"
            )
        assert res.status_code == 401
