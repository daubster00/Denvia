"""Story 6.4 — Admin 가입유형 통계 엑셀 다운로드 통합 테스트 (AC-3, AC-12)."""

from __future__ import annotations

import io
import time
from datetime import datetime
from unittest.mock import AsyncMock, MagicMock, patch

import jwt as pyjwt
import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from api.src.main import app
from api.src.models.base import get_session
from api.src.services.analytics_service import (
    EXPORT_DETAIL_LIMIT,
    SegmentRow,
    ExperienceRow,
)
from api.src.settings import settings
from api.src.services.budget_service import KST


def _make_admin_jwt() -> str:
    return pyjwt.encode(
        {
            "sub": "99",
            "aud": "denvia-admin",
            "exp": int(time.time()) + 3600,
        },
        settings.denvia_jwt_secret,
        algorithm=settings.denvia_jwt_algorithm,
    )


def _make_admin():
    user = MagicMock()
    user.id = 99
    user.email = "admin@denvia.local"
    user.role = "admin"
    user.subscription_status = "free"
    user.segment = None
    user.withdrawn_at = None
    user.must_reset_password = False
    return user


def _stub_session():
    s = MagicMock()
    s.execute = AsyncMock()
    s.commit = AsyncMock()

    async def gen():
        yield s

    return gen


def _stats() -> dict:
    return {
        "applied_filters": {"include_withdrawn": False, "include_blocked": False},
        "total": 3,
        "by_segment": [
            SegmentRow(segment="doctor", count=2, active_count=2, pro_count=1),
            SegmentRow(segment="hygienist", count=1, active_count=1, pro_count=0),
            SegmentRow(segment="student_other", count=0, active_count=0, pro_count=0),
        ],
        "by_experience": [
            ExperienceRow(segment="doctor", years_bucket="3-5", count=2),
            ExperienceRow(segment="hygienist", years_bucket="0-2", count=1),
        ],
    }


@pytest.mark.asyncio
class TestSegmentsExportAuth:
    async def test_unauth_401(self):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            res = await client.get("/api/v1/admin/analytics/segments/export")
        assert res.status_code == 401


@pytest.mark.asyncio
class TestSegmentsExport:
    async def _call(self, qs: str = ""):
        token = _make_admin_jwt()
        admin = _make_admin()
        gen = _stub_session()
        with patch(
            "api.src.deps.auth.get_user_by_id",
            new=AsyncMock(return_value=admin),
        ):
            app.dependency_overrides[get_session] = gen
            try:
                async with AsyncClient(
                    transport=ASGITransport(app=app), base_url="http://test"
                ) as client:
                    res = await client.get(
                        f"/api/v1/admin/analytics/segments/export{qs}",
                        cookies={"denvia_admin_session": token},
                    )
            finally:
                app.dependency_overrides.clear()
        return res

    async def test_basic_two_sheets(self):
        rows = [
            {
                "user_id": 1,
                "email_masked": "k**@example.com",
                "segment": "doctor",
                "segment_label": "치과의사",
                "years_of_experience": 5,
                "subscription_status": "pro",
                "created_at_kst": "2026-04-01T09:00",
            },
        ]
        with patch(
            "api.src.routers.admin.analytics.get_segment_stats",
            new=AsyncMock(return_value=_stats()),
        ), patch(
            "api.src.routers.admin.analytics.get_segment_export_rows",
            new=AsyncMock(return_value=(rows, False)),
        ):
            res = await self._call()
        assert res.status_code == 200
        assert (
            res.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Validate xlsx structure
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        assert "Summary" in wb.sheetnames
        assert "Detail" in wb.sheetnames
        det = wb["Detail"]
        # Detail header (no truncated banner here)
        det_headers = [c.value for c in det[1]]
        assert det_headers[:3] == ["user_id", "email_masked", "segment"]
        # First data row
        assert det[2][0].value == 1
        assert det[2][1].value == "k**@example.com"

    async def test_filename_kst_today(self):
        with patch(
            "api.src.routers.admin.analytics.get_segment_stats",
            new=AsyncMock(return_value=_stats()),
        ), patch(
            "api.src.routers.admin.analytics.get_segment_export_rows",
            new=AsyncMock(return_value=([], False)),
        ):
            res = await self._call()
        disp = res.headers.get("content-disposition", "")
        today_kst = datetime.now(KST).date().isoformat()
        assert f'filename="segments_{today_kst}.xlsx"' in disp

    async def test_truncated_header_set(self):
        # 5,001 행 — limit 초과를 시뮬레이션하기 위해 service mock에서 truncated=True 반환
        rows = [
            {
                "user_id": i,
                "email_masked": f"u{i}**@x.com",
                "segment": "doctor",
                "segment_label": "치과의사",
                "years_of_experience": 0,
                "subscription_status": "free",
                "created_at_kst": "2026-04-01T09:00",
            }
            for i in range(1, 11)
        ]
        with patch(
            "api.src.routers.admin.analytics.get_segment_stats",
            new=AsyncMock(return_value=_stats()),
        ), patch(
            "api.src.routers.admin.analytics.get_segment_export_rows",
            new=AsyncMock(return_value=(rows, True)),
        ):
            res = await self._call()
        assert res.status_code == 200
        assert res.headers.get("X-Truncated") == "true"
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        det = wb["Detail"]
        # 첫 행에 "5000행으로 제한됨" 캡션
        assert str(det[1][0].value).startswith(f"※ {EXPORT_DETAIL_LIMIT}행")
