"""Story 5.2 / 5.3 / 5.4 — 사용자별 토큰·비용 분석 + 가입자/구독 분포 + 피드백 분석 API."""

import io
from datetime import datetime, timedelta, date
from decimal import Decimal
from typing import Literal

import structlog
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import or_, select, func
from sqlalchemy.ext.asyncio import AsyncSession

from redis.asyncio import Redis as AsyncRedis

from api.src.deps.auth import require_admin, require_admin_page
from api.src.deps.redis import get_redis_runtime
from api.src.models.base import get_session
from api.src.models.qa_log import QALog
from api.src.models.user import User
from api.src.services.analytics_service import (
    ALLOWED_SERIES_MONTHS,
    EXPORT_DETAIL_LIMIT,
    EXPORT_DETAIL_LIMIT_REVENUE,
    QUESTIONS_EXPORT_LIMIT,
    QuestionsSort,
    SEGMENT_LABELS_KR,
    Unit,
    YEAR_MONTH_RE,
    _default_window,
    _feedback_default_window,
    _kst_datetime,
    get_feedback_export_rows,
    get_feedback_items,
    get_feedback_items_total,
    get_feedback_series,
    get_feedback_summary,
    get_questions_export_rows,
    get_questions_items,
    get_questions_items_total,
    get_questions_summary,
    get_revenue_variance_export_rows,
    get_revenue_variance_month,
    get_revenue_variance_series,
    get_segment_export_rows,
    get_segment_stats,
    get_access_buckets,
    get_signups_buckets,
    get_subscriber_counts,
)
from api.src.services import runtime_config_service
from api.src.services.budget_service import KST, kst_month_bounds
from api.src.services.finance_service import _excel_safe_cell

logger = structlog.get_logger(__name__)

router = APIRouter(
    prefix="/admin/analytics",
    tags=["admin-analytics"],
    dependencies=[Depends(require_admin_page("/admin"))],
)


class UserTokensRow(BaseModel):
    user_id: int
    email: str
    segment: str | None
    total_input_tokens: int
    total_output_tokens: int
    total_cost_usd: Decimal
    # 전체 시스템 KRW 통일 — UI 표시용 환산 보조 필드.
    total_cost_krw: int
    avg_cost_per_question_krw: int
    question_count: int
    avg_cost_per_question: Decimal


class UserTokensListResponse(BaseModel):
    items: list[UserTokensRow]
    page: int
    per_page: int
    total: int
    range: str
    year_month: str | None
    # 응답 시점에 적용된 USD→KRW 환율 (보조 표시용).
    usd_to_krw: int


@router.get("/user-tokens", response_model=UserTokensListResponse)
async def user_tokens(
    range: Literal["day", "month", "year"] = Query("month"),
    year_month: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
    redis_runtime: AsyncRedis = Depends(get_redis_runtime),
) -> UserTokensListResponse:
    start, end, ym = _resolve_window(range, year_month, from_, to)

    base = (
        select(
            QALog.user_id,
            func.coalesce(func.sum(QALog.input_tokens), 0).label("in_t"),
            func.coalesce(func.sum(QALog.output_tokens), 0).label("out_t"),
            func.coalesce(func.sum(QALog.cost_usd), Decimal("0")).label("cost"),
            func.count(QALog.id).label("q_cnt"),
        )
        .join(User, User.id == QALog.user_id)
        .where(
            QALog.created_at >= start,
            QALog.created_at < end,
            QALog.user_id.is_not(None),
            User.role != "admin",
        )
        .group_by(QALog.user_id)
    )

    total = (await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar_one()

    rows = (await db.execute(
        base.order_by(func.coalesce(func.sum(QALog.cost_usd), 0).desc())
        .offset((page - 1) * per_page).limit(per_page)
    )).all()

    user_ids = [r.user_id for r in rows]
    users_by_id: dict[int, User] = {}
    if user_ids:
        users = (await db.execute(
            select(User).where(User.id.in_(user_ids))
        )).scalars().all()
        users_by_id = {u.id: u for u in users}

    usd_to_krw = await runtime_config_service.get_usd_to_krw(redis_runtime)
    rate_dec = Decimal(usd_to_krw)

    items: list[UserTokensRow] = []
    for r in rows:
        u = users_by_id.get(r.user_id)
        email = u.email if u else f"user#{r.user_id}"
        if u and getattr(u, "withdrawn_at", None) is not None:
            email = f"{email} (탈퇴)"
        cost_dec = Decimal(str(r.cost))
        avg = (cost_dec / r.q_cnt) if r.q_cnt > 0 else Decimal("0")
        items.append(UserTokensRow(
            user_id=r.user_id,
            email=email,
            segment=(u.segment if u else None),
            total_input_tokens=int(r.in_t),
            total_output_tokens=int(r.out_t),
            total_cost_usd=cost_dec,
            total_cost_krw=int((cost_dec * rate_dec).quantize(Decimal("1"))),
            avg_cost_per_question_krw=int((avg * rate_dec).quantize(Decimal("1"))),
            question_count=int(r.q_cnt),
            avg_cost_per_question=avg,
        ))

    logger.info(
        "admin.dashboard.token_breakdown.viewed",
        actor_user_id=actor.id,
        range=range,
        year_month=ym,
        page=page,
    )

    return UserTokensListResponse(
        items=items,
        page=page,
        per_page=per_page,
        total=total,
        range=range,
        year_month=(ym if range == "month" else None),
        usd_to_krw=usd_to_krw,
    )


# ---------------------------------------------------------------------------
# Story 5.3: 가입자 추세 / 구독 분포
# ---------------------------------------------------------------------------


class SignupsBucketResponse(BaseModel):
    bucket_start: str  # YYYY-MM-DD (KST)
    cumulative: int
    active: int
    withdrawn: int
    new_signups: int


class SignupsResponse(BaseModel):
    unit: Unit
    from_: str = Field(alias="from")
    to: str
    buckets: list[SignupsBucketResponse]

    model_config = {"populate_by_name": True}


class PendingCancellation(BaseModel):
    """해지 예약 구독 — current_period_end 도래 시 종료 예정."""

    user_id: int
    email_masked: str
    canceled_at: datetime          # 사용자가 해지 신청한 시각
    current_period_end: datetime   # 유료 종료 예정 시각


class SubscribersResponse(BaseModel):
    as_of: str  # ISO-8601 KST
    free_count: int
    pro_count: int
    blocked_count: int
    withdrawn_count: int
    pending_cancellation_count: int
    pending_cancellations: list[PendingCancellation]


@router.get("/signups", response_model=SignupsResponse, response_model_by_alias=True)
async def signups(
    response: Response,
    unit: Unit = Query("month"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> SignupsResponse:
    buckets, resolved_from, resolved_to = await get_signups_buckets(db, unit, from_, to)
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.signups.viewed",
        actor_user_id=actor.id,
        unit=unit,
        from_=resolved_from.isoformat(),
        to=resolved_to.isoformat(),
        bucket_count=len(buckets),
    )
    return SignupsResponse(
        unit=unit,
        from_=resolved_from.isoformat(),
        to=resolved_to.isoformat(),
        buckets=[
            SignupsBucketResponse(
                bucket_start=b.bucket_start.isoformat(),
                cumulative=b.cumulative,
                active=b.active,
                withdrawn=b.withdrawn,
                new_signups=b.new_signups,
            )
            for b in buckets
        ],
    )


@router.get("/subscribers", response_model=SubscribersResponse)
async def subscribers(
    response: Response,
    as_of: Literal["now"] = Query("now"),  # 미래 확장 자리
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> SubscribersResponse:
    counts = await get_subscriber_counts(db)
    now_kst = datetime.now(KST).isoformat()
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.subscribers.viewed",
        actor_user_id=actor.id,
        free=counts["free_count"],
        pro=counts["pro_count"],
        blocked=counts["blocked_count"],
        withdrawn=counts["withdrawn_count"],
        pending_cancellation=counts["pending_cancellation_count"],
    )
    return SubscribersResponse(as_of=now_kst, **counts)


# =============================================================================
# 접속 통계 — login_events 기반 일/주/월/년 집계
# =============================================================================


class AccessBucketResponse(BaseModel):
    bucket_start: str  # YYYY-MM-DD (KST)
    visitors: int      # 고유 접속자 수
    visits: int        # 접속 횟수


class AccessResponse(BaseModel):
    unit: Unit
    from_: str = Field(alias="from")
    to: str
    total_visitors: int
    total_visits: int
    buckets: list[AccessBucketResponse]

    model_config = {"populate_by_name": True}


@router.get("/access", response_model=AccessResponse, response_model_by_alias=True)
async def access_stats(
    response: Response,
    unit: Unit = Query("day"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> AccessResponse:
    summary = await get_access_buckets(db, unit, from_, to)
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.access.viewed",
        actor_user_id=actor.id,
        unit=unit,
        from_=summary.from_.isoformat(),
        to=summary.to.isoformat(),
        bucket_count=len(summary.buckets),
        total_visitors=summary.total_visitors,
        total_visits=summary.total_visits,
    )
    return AccessResponse(
        unit=unit,
        from_=summary.from_.isoformat(),
        to=summary.to.isoformat(),
        total_visitors=summary.total_visitors,
        total_visits=summary.total_visits,
        buckets=[
            AccessBucketResponse(
                bucket_start=b.bucket_start.isoformat(),
                visitors=b.visitors,
                visits=b.visits,
            )
            for b in summary.buckets
        ],
    )


# =============================================================================
# Story 5.4 — 피드백 분석
# =============================================================================

RatingFilterParam = Literal["good", "bad", "all"]


class FeedbackSummary(BaseModel):
    good_count: int
    bad_count: int
    good_ratio: float | None


class FeedbackSeriesItem(BaseModel):
    bucket_start: str
    good: int
    bad: int


class FeedbackItem(BaseModel):
    qa_log_id: int
    question_text: str
    answer_text: str | None
    rating: str
    segment: str | None
    user_id: int | None = None
    email: str | None = None
    created_at: str


class FeedbackResponse(BaseModel):
    unit: str
    from_: str = Field(alias="from")
    to: str
    rating_filter: str
    summary: FeedbackSummary
    series: list[FeedbackSeriesItem]
    items: list[FeedbackItem]
    page: int
    per_page: int
    total: int

    model_config = {"populate_by_name": True}


def _resolve_feedback_window(
    unit: Literal["day", "week", "month"],
    from_: date | None,
    to: date | None,
) -> tuple[datetime, datetime, date, date]:
    """KST half-open interval 반환: [start_kst, end_exclusive_kst), 사용자 from/to date."""
    if from_ is None and to is None:
        f, t = _feedback_default_window(unit)
    elif from_ is None:
        f, _ = _feedback_default_window(unit)
        t = to  # type: ignore[assignment]
    elif to is None:
        f = from_
        _, t = _feedback_default_window(unit)
    else:
        f, t = from_, to
    start_kst = _kst_datetime(f)
    end_exclusive_kst = _kst_datetime(t) + timedelta(days=1)
    return start_kst, end_exclusive_kst, f, t


@router.get("/feedback", response_model=FeedbackResponse, response_model_by_alias=True)
async def feedback(
    response: Response,
    unit: Literal["day", "week", "month"] = Query("month"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    rating_filter: RatingFilterParam = Query("all"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    q: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> FeedbackResponse:
    q_like = f"%{q}%" if q else None
    start_kst, end_exclusive_kst, from_date, to_date = _resolve_feedback_window(
        unit, from_, to
    )

    summary_data, series_data, total = await _gather_feedback(
        db, unit, start_kst, end_exclusive_kst, rating_filter, page, per_page, q_like
    )
    items_data = await get_feedback_items(
        db, start_kst, end_exclusive_kst, rating_filter, page, per_page, q_like
    )

    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.feedback.viewed",
        actor_user_id=actor.id,
        unit=unit,
        rating_filter=rating_filter,
        page=page,
        total=total,
    )

    return FeedbackResponse(
        unit=unit,
        from_=from_date.isoformat(),
        to=to_date.isoformat(),
        rating_filter=rating_filter,
        summary=FeedbackSummary(**summary_data),
        series=[FeedbackSeriesItem(**s) for s in series_data],
        items=[FeedbackItem(**item) for item in items_data],
        page=page,
        per_page=per_page,
        total=total,
    )


async def _gather_feedback(
    db: AsyncSession,
    unit: Literal["day", "week", "month"],
    start_kst: datetime,
    end_exclusive_kst: datetime,
    rating_filter: RatingFilterParam,
    page: int,
    per_page: int,
    q_like: str | None,
) -> tuple[dict, list, int]:
    summary_data = await get_feedback_summary(db, start_kst, end_exclusive_kst, q_like)
    series_data = await get_feedback_series(db, unit, start_kst, end_exclusive_kst, q_like)
    total = await get_feedback_items_total(
        db, start_kst, end_exclusive_kst, rating_filter, q_like
    )
    return summary_data, series_data, total


class FeedbackRetrievedDocItem(BaseModel):
    """qa_logs.retrieved_docs JSONB 항목 — 관리자 감사용 top-k 단일 문서."""

    page_content: str = ""
    metadata: dict = Field(default_factory=dict)


class FeedbackDetail(BaseModel):
    """피드백 단건 상세 — Drawer '상세보기' 응답.

    관리자 감사용으로만 노출되며, SSE 사용자 응답에는 등장하지 않는
    ``normalized_query``, ``retrieved_docs``(top-k)를 포함한다.
    비회원 피드백도 조회 가능하므로 user_id 교차 검증 없이 qa_log_id 단독으로 조회한다.
    """

    qa_log_id: int
    question_text: str
    normalized_query: str | None = None
    retrieved_docs: list[FeedbackRetrievedDocItem] = Field(default_factory=list)
    answer_text: str | None = None
    rule_matched: bool = False
    status: str | None = None
    created_at: str


@router.get("/feedback/export")
async def feedback_export(
    unit: Literal["day", "week", "month"] = Query("month"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    rating_filter: RatingFilterParam = Query("all"),
    q: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    import openpyxl

    q_like = f"%{q}%" if q else None
    start_kst, end_exclusive_kst, from_date, to_date = _resolve_feedback_window(
        unit, from_, to
    )
    rows, truncated = await get_feedback_export_rows(
        db, start_kst, end_exclusive_kst, rating_filter, q_like
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "피드백"
    ws.append([
        "qa_log_id",
        "질문",
        "답변",
        "피드백",
        "계정(이메일)",
        "user_id",
        "가입유형",
        "제출일시(KST)",
    ])
    for row in rows:
        ws.append([
            row["qa_log_id"],
            row["question_text"],
            row["answer_text"] or "",
            row["rating"],
            row.get("email") or "",
            row.get("user_id") if row.get("user_id") is not None else "",
            row["segment"] or "",
            row["created_at_kst"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"feedback_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    logger.info(
        "admin.analytics.feedback.exported",
        actor_user_id=actor.id,
        row_count=len(rows),
    )

    headers: dict[str, str] = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    if truncated:
        headers["X-Truncated"] = "true"

    return StreamingResponse(buf, media_type=content_type, headers=headers)


@router.get("/feedback/{qa_log_id}", response_model=FeedbackDetail)
async def feedback_detail(
    qa_log_id: int,
    response: Response,
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> FeedbackDetail:
    """피드백 단건 상세 — top-k 검색 문서 + 동의어 치환 쿼리 포함.

    /feedback Drawer '상세보기'에서 호출. 비회원 피드백도 포함하므로
    user_id 교차 검증 없이 qa_log_id 단독으로 조회한다 (admin 전용).
    """
    row = (
        await db.execute(select(QALog).where(QALog.id == qa_log_id))
    ).scalar_one_or_none()
    if row is None:
        raise HTTPException(
            status_code=404,
            detail={
                "code": "ADMIN_FEEDBACK_QA_LOG_NOT_FOUND",
                "message": "질의 기록을 찾을 수 없습니다.",
            },
        )

    raw_docs = row.retrieved_docs or []
    docs: list[FeedbackRetrievedDocItem] = []
    for entry in raw_docs:
        if not isinstance(entry, dict):
            continue
        docs.append(
            FeedbackRetrievedDocItem(
                page_content=str(entry.get("page_content", "")),
                metadata=entry.get("metadata") or {},
            )
        )

    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.feedback.detail_viewed",
        actor_user_id=actor.id,
        qa_log_id=qa_log_id,
    )

    return FeedbackDetail(
        qa_log_id=row.id,
        question_text=row.question_text or "",
        normalized_query=row.normalized_query,
        retrieved_docs=docs,
        answer_text=row.answer_text,
        rule_matched=bool(row.rule_matched),
        status=row.status,
        created_at=row.created_at.isoformat() if row.created_at else "",
    )


# =============================================================================
# 질문 분석 — 일/주/월/년 합산 + 정렬·페이지 + 엑셀
# =============================================================================

_QUESTIONS_SEGMENT_LABELS = {
    "doctor": "치과의사",
    "hygienist": "치과위생사",
    "student_other": "학생/기타",
}


def _questions_answer_fallback(status: str | None) -> str:
    if status == "in_progress":
        return "(스트리밍 진행 중 — 응답 미저장)"
    if status == "error":
        return "(응답 생성 실패)"
    return "(응답 없음)"


class QuestionsBucketResponse(BaseModel):
    bucket_start: str  # YYYY-MM-DD (KST)
    count: int


class QuestionsItem(BaseModel):
    qa_log_id: int
    question_text: str
    answer_text: str | None
    input_tokens: int
    output_tokens: int
    total_tokens: int
    cost_usd: str | None
    status: str | None
    user_id: int | None = None
    email: str | None = None
    segment: str | None = None
    created_at: str


class QuestionsResponse(BaseModel):
    unit: Unit
    sort: QuestionsSort
    from_: str = Field(alias="from")
    to: str
    total_count: int
    buckets: list[QuestionsBucketResponse]
    items: list[QuestionsItem]
    page: int
    per_page: int
    total: int

    model_config = {"populate_by_name": True}


def _resolve_questions_window(
    unit: Unit,
    from_: date | None,
    to: date | None,
) -> tuple[datetime, datetime, date, date]:
    """KST half-open interval + 사용자 from/to."""
    if from_ is None and to is None:
        f, t = _default_window(unit)
    elif from_ is None:
        f, _ = _default_window(unit)
        t = to  # type: ignore[assignment]
    elif to is None:
        f = from_
        _, t = _default_window(unit)
    else:
        f, t = from_, to
    start_kst = _kst_datetime(f)
    end_exclusive_kst = _kst_datetime(t) + timedelta(days=1)
    return start_kst, end_exclusive_kst, f, t


@router.get(
    "/questions",
    response_model=QuestionsResponse,
    response_model_by_alias=True,
)
async def questions(
    response: Response,
    unit: Unit = Query("day"),
    sort: QuestionsSort = Query("latest"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    q: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> QuestionsResponse:
    q_like = f"%{q}%" if q else None
    summary = await get_questions_summary(db, unit, from_, to)
    start_kst, end_exclusive_kst, from_date, to_date = _resolve_questions_window(
        unit, from_, to
    )
    total = await get_questions_items_total(
        db, start_kst, end_exclusive_kst, q_like
    )
    items_data = await get_questions_items(
        db, start_kst, end_exclusive_kst, sort, page, per_page, q_like
    )

    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.questions.viewed",
        actor_user_id=actor.id,
        unit=unit,
        sort=sort,
        page=page,
        total=total,
        total_count=summary.total_count,
    )

    return QuestionsResponse(
        unit=unit,
        sort=sort,
        from_=from_date.isoformat(),
        to=to_date.isoformat(),
        total_count=summary.total_count,
        buckets=[
            QuestionsBucketResponse(
                bucket_start=b.bucket_start.isoformat(),
                count=b.count,
            )
            for b in summary.buckets
        ],
        items=[QuestionsItem(**item) for item in items_data],
        page=page,
        per_page=per_page,
        total=total,
    )


@router.get("/questions/export")
async def questions_export(
    unit: Unit = Query("day"),
    sort: QuestionsSort = Query("latest"),
    from_: date | None = Query(None, alias="from"),
    to: date | None = Query(None),
    q: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    import openpyxl

    q_like = f"%{q}%" if q else None
    summary = await get_questions_summary(db, unit, from_, to)
    start_kst, end_exclusive_kst, from_date, to_date = _resolve_questions_window(
        unit, from_, to
    )
    rows, truncated = await get_questions_export_rows(
        db, start_kst, end_exclusive_kst, sort, q_like
    )

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["항목", "값"])
    ws_sum.append(["기간 (KST)", f"{from_date.isoformat()} ~ {to_date.isoformat()}"])
    ws_sum.append(["집계 단위", unit])
    ws_sum.append(["정렬", sort])
    ws_sum.append(["총 질문 수", summary.total_count])
    ws_sum.append(["행 제한 (Detail)", QUESTIONS_EXPORT_LIMIT])
    ws_sum.append(["잘림 여부", "예" if truncated else "아니오"])
    ws_sum.append([])
    ws_sum.append(["bucket_start", "질문 수"])
    for b in summary.buckets:
        ws_sum.append([b.bucket_start.isoformat(), b.count])

    ws_det = wb.create_sheet("Detail")
    if truncated:
        ws_det.append([f"※ {QUESTIONS_EXPORT_LIMIT}행으로 제한됨"])
    # 컬럼 순서·라벨을 화면 테이블과 동일하게 맞춤.
    ws_det.append(
        [
            "작성일시 (KST)",
            "계정",
            "가입유형",
            "질문",
            "답변",
            "토큰 (입력+출력)",
        ]
    )
    for r in rows:
        in_t = int(r["input_tokens"] or 0)
        out_t = int(r["output_tokens"] or 0)
        total_t = in_t + out_t
        token_cell = f"{total_t} ({in_t}+{out_t})"
        account = r["email"] or "(비회원)"
        segment_label = _QUESTIONS_SEGMENT_LABELS.get(r["segment"], r["segment"] or "—")
        answer = r["answer_text"]
        if not answer:
            answer = _questions_answer_fallback(r["status"])
        ws_det.append(
            [
                r["created_at_kst"],
                account,
                segment_label,
                r["question_text"],
                answer,
                token_cell,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"questions_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    logger.info(
        "admin.analytics.questions.exported",
        actor_user_id=actor.id,
        row_count=len(rows),
        truncated=truncated,
        unit=unit,
        sort=sort,
    )

    headers: dict[str, str] = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }
    if truncated:
        headers["X-Truncated"] = "true"

    return StreamingResponse(buf, media_type=content_type, headers=headers)


def _resolve_window(
    range_: str,
    year_month: str | None,
    from_: date | None,
    to: date | None,
) -> tuple[datetime, datetime, str | None]:
    if range_ == "month":
        if year_month is None:
            return kst_month_bounds()
        y, m = (int(p) for p in year_month.split("-"))
        start = datetime(y, m, 1, tzinfo=KST)
        end = (
            datetime(y + 1, 1, 1, tzinfo=KST)
            if m == 12
            else datetime(y, m + 1, 1, tzinfo=KST)
        )
        return start, end, year_month
    if range_ == "day":
        end_d = to or datetime.now(KST).date()
        start_d = from_ or (end_d - timedelta(days=30))
        return (
            datetime(start_d.year, start_d.month, start_d.day, tzinfo=KST),
            datetime(end_d.year, end_d.month, end_d.day, tzinfo=KST)
            + timedelta(days=1),
            None,
        )
    # year — from/to 로 특정 연도 지정 가능. 미지정 시 현재 연도.
    if from_ is not None or to is not None:
        end_d = to or datetime.now(KST).date()
        start_d = from_ or date(end_d.year, 1, 1)
        return (
            datetime(start_d.year, start_d.month, start_d.day, tzinfo=KST),
            datetime(end_d.year, end_d.month, end_d.day, tzinfo=KST)
            + timedelta(days=1),
            None,
        )
    now = datetime.now(KST)
    return (
        datetime(now.year, 1, 1, tzinfo=KST),
        datetime(now.year + 1, 1, 1, tzinfo=KST),
        None,
    )


# =============================================================================
# Story 6.4 — 가입유형 통계 + 연차 히스토그램 + 엑셀 export
# =============================================================================


class SegmentRowResponse(BaseModel):
    segment: Literal["doctor", "hygienist", "student_other"]
    count: int
    active_count: int
    pro_count: int


class ExperienceRowResponse(BaseModel):
    segment: Literal["doctor", "hygienist"]
    years_bucket: Literal["0-2", "3-5", "6-10", "11-20", "20+"]
    count: int


class SegmentsAppliedFilters(BaseModel):
    include_withdrawn: bool
    include_blocked: bool


class SegmentsResponse(BaseModel):
    as_of: str
    applied_filters: SegmentsAppliedFilters
    total: int
    by_segment: list[SegmentRowResponse]
    by_experience: list[ExperienceRowResponse]


@router.get("/segments", response_model=SegmentsResponse)
async def segments(
    response: Response,
    as_of: Literal["now"] = Query("now"),
    include_withdrawn: bool = Query(False),
    include_blocked: bool = Query(False),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> SegmentsResponse:
    stats = await get_segment_stats(
        db,
        include_withdrawn=include_withdrawn,
        include_blocked=include_blocked,
    )
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.analytics.segments.viewed",
        actor_user_id=actor.id,
        total=stats["total"],
        include_withdrawn=include_withdrawn,
        include_blocked=include_blocked,
    )
    return SegmentsResponse(
        as_of=datetime.now(KST).isoformat(),
        applied_filters=SegmentsAppliedFilters(**stats["applied_filters"]),
        total=stats["total"],
        by_segment=[SegmentRowResponse(**r.__dict__) for r in stats["by_segment"]],
        by_experience=[
            ExperienceRowResponse(**r.__dict__) for r in stats["by_experience"]
        ],
    )


@router.get("/segments/export")
async def segments_export(
    include_withdrawn: bool = Query(False),
    include_blocked: bool = Query(False),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    import openpyxl

    stats = await get_segment_stats(
        db,
        include_withdrawn=include_withdrawn,
        include_blocked=include_blocked,
    )
    rows, truncated = await get_segment_export_rows(
        db,
        include_withdrawn=include_withdrawn,
        include_blocked=include_blocked,
    )

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["가입유형 통계 — 기준", datetime.now(KST).isoformat()])
    ws_sum.append([])
    ws_sum.append(["segment", "segment_label", "count", "active_count", "pro_count"])
    for r in stats["by_segment"]:
        ws_sum.append(
            [
                r.segment,
                SEGMENT_LABELS_KR.get(r.segment, ""),
                r.count,
                r.active_count,
                r.pro_count,
            ]
        )
    ws_sum.append([])
    ws_sum.append(["segment", "segment_label", "years_bucket", "count"])
    for r in stats["by_experience"]:
        ws_sum.append(
            [
                r.segment,
                SEGMENT_LABELS_KR.get(r.segment, ""),
                r.years_bucket,
                r.count,
            ]
        )

    ws_det = wb.create_sheet("Detail")
    if truncated:
        ws_det.append([f"※ {EXPORT_DETAIL_LIMIT}행으로 제한됨"])
    ws_det.append(
        [
            "user_id",
            "email_masked",
            "segment",
            "segment_label",
            "years_of_experience",
            "subscription_status",
            "created_at_kst",
        ]
    )
    for row in rows:
        ws_det.append(
            [
                row["user_id"],
                row["email_masked"],
                row["segment"],
                row["segment_label"],
                row["years_of_experience"],
                row["subscription_status"],
                row["created_at_kst"],
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now(KST).date().isoformat()
    filename = f"segments_{today}.xlsx"
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    logger.info(
        "admin.analytics.segments.exported",
        actor_user_id=actor.id,
        row_count=len(rows),
        truncated=truncated,
    )

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    if truncated:
        headers["X-Truncated"] = "true"
    return StreamingResponse(buf, media_type=content_type, headers=headers)


# =============================================================================
# Story 5.5 — 매출 + 토큰비용 차액 (A-106) + xlsx export (A-107)
# =============================================================================


def _validate_year_month(year_month: str | None) -> str:
    """year_month 검증 + None이면 현재 KST 월로 fallback."""
    if year_month is None:
        _start, _next, ym = kst_month_bounds()
        return ym
    if not YEAR_MONTH_RE.match(year_month):
        raise HTTPException(
            status_code=422,
            detail={
                "code": "INVALID_PARAM",
                "message": "year_month는 YYYY-MM 형식이어야 합니다.",
            },
        )
    return year_month


class RevenueVarianceFilters(BaseModel):
    year_month: str
    kst_start: str
    kst_end_exclusive: str


class RevenueVarianceResponse(BaseModel):
    year_month: str
    revenue_krw: int  # gross alias — 하위호환
    gross_revenue_krw: int
    refund_krw: int
    net_revenue_krw: int
    token_cost_usd: str
    token_cost_krw: int
    usd_to_krw: int
    variance_krw: int  # net_revenue_krw - token_cost_krw
    error_count: int
    anomaly_count: int
    applied_filters: RevenueVarianceFilters


class RevenueSeriesItem(BaseModel):
    year_month: str
    revenue_krw: int  # gross alias — 하위호환
    gross_revenue_krw: int
    refund_krw: int
    net_revenue_krw: int
    token_cost_krw: int
    variance_krw: int  # net_revenue_krw - token_cost_krw


class RevenueSeriesResponse(BaseModel):
    months: int
    to: str
    from_: str = Field(alias="from")
    usd_to_krw: int
    items: list[RevenueSeriesItem]

    model_config = {"populate_by_name": True}


@router.get(
    "/revenue-variance",
    response_model=RevenueVarianceResponse,
)
async def revenue_variance(
    response: Response,
    year_month: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
    redis_runtime: AsyncRedis = Depends(get_redis_runtime),
) -> RevenueVarianceResponse:
    ym = _validate_year_month(year_month)
    data = await get_revenue_variance_month(
        db, redis_runtime=redis_runtime, year_month=ym
    )
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.finance.revenue_variance.viewed",
        actor_user_id=actor.id,
        year_month=ym,
        revenue_krw=data["revenue_krw"],
        variance_krw=data["variance_krw"],
    )
    return RevenueVarianceResponse(**data)


@router.get(
    "/revenue-variance/series",
    response_model=RevenueSeriesResponse,
    response_model_by_alias=True,
)
async def revenue_variance_series(
    response: Response,
    months: int = Query(12),
    to: str | None = Query(None),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
    redis_runtime: AsyncRedis = Depends(get_redis_runtime),
) -> RevenueSeriesResponse:
    if months not in ALLOWED_SERIES_MONTHS:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "INVALID_PARAM",
                "message": "months는 3/6/12/24 중 하나여야 합니다.",
            },
        )
    to_ym = _validate_year_month(to)
    data = await get_revenue_variance_series(
        db, redis_runtime=redis_runtime, months=months, to_year_month=to_ym
    )
    response.headers["Cache-Control"] = "no-store"
    logger.info(
        "admin.finance.revenue_variance.series.viewed",
        actor_user_id=actor.id,
        months=months,
        to=to_ym,
    )
    return RevenueSeriesResponse(
        months=data["months"],
        to=data["to"],
        usd_to_krw=data["usd_to_krw"],
        items=[RevenueSeriesItem(**i) for i in data["items"]],
        **{"from": data["from"]},
    )


@router.get("/revenue-variance/export")
async def revenue_variance_export(
    year_month: str = Query(...),
    actor: User = Depends(require_admin),
    db: AsyncSession = Depends(get_session),
    redis_runtime: AsyncRedis = Depends(get_redis_runtime),
) -> StreamingResponse:
    import openpyxl

    ym = _validate_year_month(year_month)
    summary = await get_revenue_variance_month(
        db, redis_runtime=redis_runtime, year_month=ym
    )
    rows, truncated = await get_revenue_variance_export_rows(
        db, year_month=ym, limit=EXPORT_DETAIL_LIMIT_REVENUE
    )

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["항목", "값"])
    ws_sum.append(["기간", f"{ym} (KST 1일 00:00 ~ 다음 달 1일 00:00)"])
    ws_sum.append(["당월 총매출 (KRW)", summary["gross_revenue_krw"]])
    ws_sum.append(["당월 환불액 (KRW)", summary["refund_krw"]])
    ws_sum.append(["당월 순매출 (KRW)", summary["net_revenue_krw"]])
    ws_sum.append(["당월 토큰 비용 (USD)", summary["token_cost_usd"]])
    ws_sum.append(["적용 환율 (KRW/USD)", summary["usd_to_krw"]])
    ws_sum.append(["당월 토큰 비용 (KRW)", summary["token_cost_krw"]])
    ws_sum.append(["차액 (KRW, 순매출−토큰비용)", summary["variance_krw"]])
    ws_sum.append(["결제 실패 건수", summary["error_count"]])
    ws_sum.append(["이상 이벤트 건수", summary["anomaly_count"]])
    ws_sum.append(["행 제한 (Detail)", EXPORT_DETAIL_LIMIT_REVENUE])
    ws_sum.append(["잘림 여부", "예" if truncated else "아니오"])

    ws_det = wb.create_sheet("Detail")
    if truncated:
        ws_det.append([f"※ {EXPORT_DETAIL_LIMIT_REVENUE}행으로 제한됨"])
    ws_det.append(
        [
            "payment_id",
            "charged_at_kst",
            "amount_krw",
            "user_id",
            "email_masked",
            "provider_order_id",
            "subscription_id",
            "status",
            "환불 여부",
        ]
    )
    for r in rows:
        ws_det.append(
            [
                _excel_safe_cell(r["payment_id"]),
                _excel_safe_cell(r["charged_at_kst"]),
                _excel_safe_cell(r["amount_krw"]),
                _excel_safe_cell(r["user_id"]),
                _excel_safe_cell(r["email_masked"]),
                _excel_safe_cell(r["provider_order_id"]),
                _excel_safe_cell(r["subscription_id"]),
                _excel_safe_cell(r["status"]),
                _excel_safe_cell(r["is_refunded"]),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"revenue_variance_{ym}.xlsx"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }
    if truncated:
        headers_resp["X-Truncated"] = "true"

    logger.info(
        "admin.finance.revenue_variance.exported",
        actor_user_id=actor.id,
        year_month=ym,
        row_count=len(rows),
        truncated=truncated,
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )
