"""Story 6.1 — Admin 사용자 통합 검색·상세 도메인 로직.

본 모듈은 두 가지 진입점을 제공한다:
- search_users(...): 이메일/휴대폰/카드 뒷4자리 OR-검색 + 필터 + 페이지네이션
- get_user_detail(user_id): Drawer 4 섹션(기본/결제/최근 QA/이상 이벤트) 단건 조회

설계 결정:
- OR-검색은 단일 WHERE 절(or_)로 작성. q가 비어있거나 미지정이면 OR 분기 자체를 SQL에서 제외.
- card_last4 매칭은 q 길이가 정확히 4이고 모두 숫자일 때만 EXISTS billing_keys 분기 평가
  (불필요한 join 비용 절감, NFR-P5).
- 활성 빌링키 LEFT JOIN 은 user당 1건 제약을 service 레이어에서 강제하지 않고
  중복 시 MAX(card_last4) 결정론적 선택 + structlog warn 로그(0건이면 침묵).
- total은 동일 WHERE 절을 적용한 SELECT COUNT(*) 1쿼리, items는 페이지네이션 적용된
  본 SELECT 1쿼리 — 합 2쿼리(N+1 금지).
- 정렬은 (subscription_status='blocked')::int DESC, withdrawn_at IS NULL DESC, created_at DESC.
  CASE WHEN 패턴으로 인덱스 호환성 확보.
"""

from __future__ import annotations

from datetime import date, datetime, time, timedelta
from typing import Any, Literal
from zoneinfo import ZoneInfo

import structlog
from fastapi import HTTPException
from sqlalchemy import case, exists, func, literal, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.src.models.anomaly_event import AnomalyEvent
from api.src.models.billing_key import BillingKey
from api.src.models.qa_log import QALog
from api.src.models.subscription import Subscription
from api.src.models.user import User
from api.src.schemas.admin.users import (
    RecentAnomalyEvent,
    RecentQALog,
    SubscriptionSummary,
    UserDetailResponse,
    UserSearchItem,
    UserSearchListResponse,
)

logger = structlog.get_logger(__name__)

_RECENT_QA_LIMIT = 5
_RECENT_ANOMALY_LIMIT = 3
_KST = ZoneInfo("Asia/Seoul")


def _kst_created_range(
    f: date | None, t: date | None
) -> tuple[datetime | None, datetime | None]:
    """가입일(created_at) from/to 날짜를 KST [시작, 끝+1일) datetime 범위로 변환.

    admin_support_service._kst_date_range와 동일 패턴. 종료일은 그날 끝까지
    포함되도록 다음날 00:00 KST를 exclusive upper bound로 사용한다.
    """
    start_dt: datetime | None = None
    end_dt: datetime | None = None
    if f is not None:
        start_dt = datetime.combine(f, time.min).replace(tzinfo=_KST)
    if t is not None:
        end_dt = datetime.combine(t + timedelta(days=1), time.min).replace(tzinfo=_KST)
    return start_dt, end_dt


def _is_card_last4_query(q: str) -> bool:
    """q가 카드 뒷4자리 매칭 분기를 평가할 가치가 있는지 결정.

    정확히 4자리이고 모두 숫자일 때만 billing_keys 서브쿼리를 평가한다.
    """
    return len(q) == 4 and q.isdigit()


def _build_or_clause(q: str | None) -> Any | None:
    """q OR-검색 절(이메일 ILIKE / 휴대폰 LIKE / card_last4 EXISTS)을 생성.

    q가 비어있거나 미지정이면 None을 반환해 호출자가 WHERE 절에서 제외하도록 한다.
    """
    if not q:
        return None
    pattern = f"%{q}%"
    branches = [
        func.lower(User.email).like(func.lower(pattern)),
    ]
    # 휴대폰은 NULL 가능 — LIKE 자체가 NULL을 자동 제외함
    branches.append(User.phone.like(pattern))

    if _is_card_last4_query(q):
        billing_exists = (
            select(literal(1))
            .where(
                BillingKey.user_id == User.id,
                BillingKey.is_active.is_(True),
                BillingKey.card_last4 == q,
            )
            .exists()
        )
        branches.append(billing_exists)

    return or_(*branches)


async def _resolve_active_billing_keys(
    db: AsyncSession, user_ids: list[int]
) -> dict[int, tuple[str | None, str | None]]:
    """해당 user_id 집합의 활성 빌링키를 user_id → (card_last4, card_company) dict로 반환.

    한 user에 활성 빌링키가 2건 이상이면 MAX(card_last4) 결정론적 선택 + warn 로그.
    """
    if not user_ids:
        return {}
    rows = (
        await db.execute(
            select(BillingKey.user_id, BillingKey.card_last4, BillingKey.card_company)
            .where(BillingKey.user_id.in_(user_ids), BillingKey.is_active.is_(True))
            .order_by(BillingKey.user_id, BillingKey.card_last4)
        )
    ).all()

    grouped: dict[int, list[tuple[str | None, str | None]]] = {}
    for row in rows:
        grouped.setdefault(row.user_id, []).append((row.card_last4, row.card_company))

    result: dict[int, tuple[str | None, str | None]] = {}
    for uid, entries in grouped.items():
        if len(entries) > 1:
            logger.warning(
                "admin.users.duplicate_active_billing_key",
                target_user_id=uid,
                active_count=len(entries),
            )
            entries.sort(key=lambda t: (t[0] or ""), reverse=True)
        result[uid] = entries[0]
    return result


def _serialize_user(
    user: User, billing: tuple[str | None, str | None] | None
) -> UserSearchItem:
    """User ORM + 빌링키 튜플을 응답 schema로 직렬화."""
    card_last4, card_company = billing if billing else (None, None)
    is_blocked = user.subscription_status == "blocked"
    return UserSearchItem(
        user_id=user.id,
        email=user.email,
        phone=user.phone,
        segment=user.segment,
        years_of_experience=user.years_of_experience,
        subscription_status=user.subscription_status,  # type: ignore[arg-type]
        is_blocked=is_blocked,
        block_until=None,  # Story 6.2 채움
        daily_quota_override=user.daily_quota_override,
        free_delay_override=(
            float(user.free_delay_override) if user.free_delay_override is not None else None
        ),
        anomaly_throttled_at=user.anomaly_throttled_at,
        created_at=user.created_at,
        last_login_at=user.last_login_at,
        withdrawn_at=user.withdrawn_at,
        pro_since=None,  # Story 6.2-followup
        card_last4=card_last4,
        card_company=card_company,
        name=user.name,
        birthdate=user.birthdate,
        gender=user.gender,  # type: ignore[arg-type]
        postcode=user.postcode,
        address_road=user.address_road,
        address_detail=user.address_detail,
        marketing_consent_at=user.marketing_consent_at,
    )


async def search_users(
    db: AsyncSession,
    *,
    q: str | None = None,
    segment: Literal["doctor", "hygienist", "student_other"] | None = None,
    subscription_status: Literal["free", "pro", "blocked"] | None = None,
    blocked: bool | None = None,
    withdrawn: bool | None = None,
    created_from: date | None = None,
    created_to: date | None = None,
    page: int = 1,
    per_page: int = 20,
) -> UserSearchListResponse:
    """이메일/휴대폰/카드 뒷4자리 OR-검색 + 필터 + 페이지네이션.

    blocked=True 면 subscription_status='blocked'만, blocked=False 면 그 외만,
    None이면 모든 상태를 포함한다.

    withdrawn=True 면 withdrawn_at IS NOT NULL 만, withdrawn=False 면 IS NULL 만,
    None이면 탈퇴 여부 무관. 대시보드 구독 현황의 무료/Pro/차단 카운트는
    withdrawn_at IS NULL 을 강제하므로, 대시보드에서 진입한 경우 withdrawn=False
    를 함께 전달해야 동일 카운트가 나온다.

    created_from/created_to 는 KST 기준 가입일 범위 필터(양 끝일 포함). 한쪽만
    지정해도 동작한다.
    """
    # 1. WHERE 절 구성 — None인 필터는 SQL에서 제외 (불필요한 분기 회피)
    conditions: list[Any] = [User.role != "admin"]  # 관리자 계정 제외
    if segment is not None:
        conditions.append(User.segment == segment)
    if subscription_status is not None:
        conditions.append(User.subscription_status == subscription_status)
    if blocked is True:
        conditions.append(User.subscription_status == "blocked")
    elif blocked is False:
        conditions.append(User.subscription_status != "blocked")
    if withdrawn is True:
        conditions.append(User.withdrawn_at.is_not(None))
    elif withdrawn is False:
        conditions.append(User.withdrawn_at.is_(None))

    start_dt, end_dt = _kst_created_range(created_from, created_to)
    if start_dt is not None:
        conditions.append(User.created_at >= start_dt)
    if end_dt is not None:
        conditions.append(User.created_at < end_dt)

    or_clause = _build_or_clause((q or "").strip() or None)
    if or_clause is not None:
        conditions.append(or_clause)

    # 2. total 카운트 1쿼리 (페이지네이션 메타)
    total_stmt = select(func.count()).select_from(User)
    if conditions:
        total_stmt = total_stmt.where(*conditions)
    total = (await db.execute(total_stmt)).scalar_one()

    # 3. items 페이지 SELECT 1쿼리 — 정렬 키 CASE WHEN으로 인덱스 호환
    blocked_first = case((User.subscription_status == "blocked", 0), else_=1)
    withdrawn_last = case((User.withdrawn_at.is_(None), 0), else_=1)

    items_stmt = (
        select(User)
        .order_by(blocked_first.asc(), withdrawn_last.asc(), User.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    if conditions:
        items_stmt = items_stmt.where(*conditions)

    user_rows = (await db.execute(items_stmt)).scalars().all()

    # 4. 활성 빌링키 1쿼리 — user_id 집합으로 IN 쿼리 (N+1 방지)
    user_ids = [u.id for u in user_rows]
    billing_map = await _resolve_active_billing_keys(db, user_ids)

    items = [_serialize_user(u, billing_map.get(u.id)) for u in user_rows]

    return UserSearchListResponse(
        items=items, page=page, per_page=per_page, total=int(total)
    )


async def get_user_detail(db: AsyncSession, user_id: int) -> UserDetailResponse:
    """단건 사용자 상세 — 기본 정보 + 결제 요약 + 최근 QA 5 + 이상 이벤트 3.

    존재하지 않으면 404 ADMIN_USER_NOT_FOUND.
    """
    user = (
        await db.execute(select(User).where(User.id == user_id))
    ).scalar_one_or_none()
    if user is None:
        raise HTTPException(
            status_code=404,
            detail={
                "code": "ADMIN_USER_NOT_FOUND",
                "message": "사용자를 찾을 수 없습니다.",
            },
        )

    billing_map = await _resolve_active_billing_keys(db, [user.id])
    billing = billing_map.get(user.id)
    user_item = _serialize_user(user, billing)

    # subscriptions: status='active'인 단일 행 (없으면 NULL 자리)
    subscription = (
        await db.execute(
            select(Subscription)
            .where(Subscription.user_id == user_id, Subscription.status == "active")
            .limit(1)
        )
    ).scalar_one_or_none()

    subscription_summary = SubscriptionSummary(
        current_status=user.subscription_status,  # type: ignore[arg-type]
        billing_key_active=billing is not None,
        card_last4=billing[0] if billing else None,
        card_company=billing[1] if billing else None,
        subscription_started_at=subscription.started_at if subscription else None,
        next_charge_at=subscription.next_charge_at if subscription else None,
    )

    # 최근 QA 5건
    qa_rows = (
        await db.execute(
            select(QALog)
            .where(QALog.user_id == user_id)
            .order_by(QALog.created_at.desc())
            .limit(_RECENT_QA_LIMIT)
        )
    ).scalars().all()
    recent_qa = [
        RecentQALog(
            qa_log_id=row.id,
            question_excerpt=row.question_text or "",
            answer_excerpt=row.answer_text if row.answer_text else None,
            input_tokens=row.input_tokens,
            output_tokens=row.output_tokens,
            cost_usd=row.cost_usd,
            status=row.status,
            created_at=row.created_at,
        )
        for row in qa_rows
    ]

    # 최근 이상 이벤트 3건
    anomaly_rows = (
        await db.execute(
            select(AnomalyEvent)
            .where(AnomalyEvent.target_user_id == user_id)
            .order_by(AnomalyEvent.created_at.desc())
            .limit(_RECENT_ANOMALY_LIMIT)
        )
    ).scalars().all()
    recent_anomaly = [
        RecentAnomalyEvent(
            id=row.id,
            type=row.type,
            ip=row.ip,
            status=row.status,
            created_at=row.created_at,
        )
        for row in anomaly_rows
    ]

    return UserDetailResponse(
        user=user_item,
        subscription_summary=subscription_summary,
        recent_qa=recent_qa,
        recent_anomaly_events=recent_anomaly,
    )


# ── 고객 기본정보 엑셀 내보내기 ────────────────────────────────────────────────
# 관리자 페이지 "고객관리" 헤더의 [엑셀 다운로드] 버튼이 호출한다.
# 현재 검색 필터를 그대로 적용해 같은 결과를 1장의 시트로 내려준다.
# PII 보호를 위해 응답 바이트는 메모리에 stream BytesIO 로만 보관한다.

_EXPORT_HEADERS: list[tuple[str, str]] = [
    ("user_id", "고객번호"),
    ("email", "이메일"),
    ("phone", "휴대폰"),
    ("name", "이름"),
    ("gender", "성별"),
    ("birthdate", "생년월일"),
    ("postcode", "우편번호"),
    ("address_road", "도로명주소"),
    ("address_detail", "상세주소"),
    ("segment", "가입유형"),
    ("years_of_experience", "연차"),
    ("subscription_status", "구독상태"),
    ("marketing_consent", "마케팅 수신동의"),
    ("created_at", "가입일"),
    ("last_login_at", "최근 로그인"),
    ("withdrawn_at", "탈퇴일"),
]

_SEGMENT_LABEL_KO = {
    "doctor": "치과의사",
    "hygienist": "치과위생사",
    "student_other": "학생/기타",
}

_SUBSCRIPTION_LABEL_KO = {
    "free": "무료",
    "pro": "Pro",
    "blocked": "차단",
}

_GENDER_LABEL_KO = {"male": "남", "female": "여"}


def _format_kst_datetime(value: datetime | None) -> str:
    """엑셀 셀 출력용 KST 시각 포매팅 — None이면 빈 문자열."""
    if value is None:
        return ""
    try:
        return value.astimezone(_KST).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return value.isoformat()


def _user_row_for_export(user: User) -> list[Any]:
    """User ORM 1행 → 엑셀 cell 값 리스트(헤더 순서 일치)."""
    return [
        user.id,
        user.email or "",
        user.phone or "",
        user.name or "",
        _GENDER_LABEL_KO.get(user.gender or "", user.gender or ""),
        user.birthdate.strftime("%Y-%m-%d") if user.birthdate else "",
        user.postcode or "",
        user.address_road or "",
        user.address_detail or "",
        _SEGMENT_LABEL_KO.get(user.segment or "", user.segment or ""),
        user.years_of_experience if user.years_of_experience is not None else "",
        _SUBSCRIPTION_LABEL_KO.get(
            user.subscription_status, user.subscription_status
        ),
        "동의" if user.marketing_consent_at is not None else "미동의",
        _format_kst_datetime(user.created_at),
        _format_kst_datetime(user.last_login_at),
        _format_kst_datetime(user.withdrawn_at),
    ]


async def export_users_xlsx(
    db: AsyncSession,
    *,
    q: str | None = None,
    segment: Literal["doctor", "hygienist", "student_other"] | None = None,
    subscription_status: Literal["free", "pro", "blocked"] | None = None,
    blocked: bool | None = None,
    withdrawn: bool | None = None,
    created_from: date | None = None,
    created_to: date | None = None,
) -> tuple[bytes, int]:
    """검색 조건과 동일한 결과를 xlsx 바이트로 직렬화한다.

    반환값: (xlsx 바이트, 행 수). 페이지네이션은 적용하지 않고 조건에 맞는 전체를
    한 번에 내려준다(관리자 운영용 일괄 내려받기). 너무 많을 경우 검색 필터로
    줄이도록 안내(프론트). 정렬 키는 search_users 와 동일하게 유지한다.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    conditions: list[Any] = [User.role != "admin"]
    if segment is not None:
        conditions.append(User.segment == segment)
    if subscription_status is not None:
        conditions.append(User.subscription_status == subscription_status)
    if blocked is True:
        conditions.append(User.subscription_status == "blocked")
    elif blocked is False:
        conditions.append(User.subscription_status != "blocked")
    if withdrawn is True:
        conditions.append(User.withdrawn_at.is_not(None))
    elif withdrawn is False:
        conditions.append(User.withdrawn_at.is_(None))

    start_dt, end_dt = _kst_created_range(created_from, created_to)
    if start_dt is not None:
        conditions.append(User.created_at >= start_dt)
    if end_dt is not None:
        conditions.append(User.created_at < end_dt)

    or_clause = _build_or_clause((q or "").strip() or None)
    if or_clause is not None:
        conditions.append(or_clause)

    blocked_first = case((User.subscription_status == "blocked", 0), else_=1)
    withdrawn_last = case((User.withdrawn_at.is_(None), 0), else_=1)

    stmt = (
        select(User)
        .where(*conditions)
        .order_by(blocked_first.asc(), withdrawn_last.asc(), User.created_at.desc())
    )
    rows = (await db.execute(stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "고객 기본정보"

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(bold=True, color="0F172A")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, (_, label) in enumerate(_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for r, user in enumerate(rows, start=2):
        for c, value in enumerate(_user_row_for_export(user), start=1):
            ws.cell(row=r, column=c, value=value)

    # 헤더 폭 자동 — 한글은 글자당 약 1.6배 폭 가산.
    for col_idx, (_, label) in enumerate(_EXPORT_HEADERS, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = max(12, len(label) * 2 + 2)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows)


async def count_registered_users(db: AsyncSession) -> int:
    """가입된 일반 사용자(role='user', 관리자 제외) 총수.

    검색·필터와 무관한 절대값 — 고객관리 상단 가입자 수 배지용.
    탈퇴자(withdrawn_at) 포함: '가입자 수' 1차 요구라 단순 총계로 센다.
    """
    stmt = select(func.count()).select_from(User).where(User.role == "user")
    return int((await db.execute(stmt)).scalar_one())


__all__ = [
    "search_users",
    "get_user_detail",
    "export_users_xlsx",
    "count_registered_users",
]
